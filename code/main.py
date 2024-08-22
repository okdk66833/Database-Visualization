#================모듈================#
from openpyxl import load_workbook
import os
import ntpath
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
from random import *
#import tkinter.ttk
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

from matplotlib import font_manager, rc
plt.rcParams['axes.unicode_minus'] = False
f_path = "C:/Windows/Fonts/malgun.ttf"
font_name=font_manager.FontProperties(fname=f_path).get_name()
rc('font', family=font_name)
#================root================#
root=Tk()
root.geometry("325x124")
root.config(bg="#fffeff")
root.title("HU-DE")
#================변수================#
Lists=[]
Colors=[]
data=[]
patch=[]
x=[]
y=[]
graph=[]
xname=""
yname=""
unit=""
title=""
color=[]

Selx=0          #스케터 x
Sely=0          #스케터 y
selx,sely=0,0
FileLocate=""
FileName=""
wb=""
ws=""
TableN=""
NN=""
DFontSize=15    #일반 폰트 크기
TFontSize=25    #타이틀 폰트 크기
SFontSize=17    #서브 타이틀 폰트 크기
#================함수================#
def hexrandom():
    r,g,b=randint(0,255),randint(0,255),randint(0,255)
    return f'#{r:02x}{g:02x}{b:02x}'
    
def TypeA():
    plt.cla()
    for i in range(NN):
        color.append(hexrandom())
        plt.plot(x,List[i],color=color[i],
             marker='o', markerfacecolor=color[i],
             markersize=6)    
    plt.legend(y)    
    plt.xlabel(xname)
    plt.ylabel(unit)
    plt.title(title)
    plt.show()

def TypeB():
    plt.cla()
    global Lists, Colors, data, patch
    patch=[]
    data=[]
    Lists=[]
    Colors=[]
    
    for i in range(NN):
        color.append(hexrandom())
        patch.append(mpatches.Patch(color=color[i],label=y[i]))
    
    for i in range(NN):
        for j in range(TableN):
            data.append((j,List[i][j],y[i]))
    data.sort()
    for i in range(NN):
        Lists.clear()
        Colors.clear()
        for j in range(TableN):
            for k in range(NN):
                if y[k]==data[NN*(j+1)-i-1][2]:
                    Colors.append(color[k])
            Lists.append(data[NN*(j+1)-i-1][1])
        plt.bar(x,Lists,color=Colors)
    plt.legend(handles=patch)
    plt.xlabel(xname)
    plt.ylabel(unit)
    plt.title(title)
    plt.show()

def TypeC():
    pass
    
def TypeC_2():
    global Selx,Sely,x,y
    global selx,sely
    selx,sely=Selx,Sely
    def Sel1(a):
        global selx,sely,x,y
        for i in range(TableN):
            q=list(Clistbox01.curselection())
            if x[i]==x[q[0]]:
                selx=i
        Clabel02.config(text=f"x축 : {x[selx]}")
    def Sel2(a):
        global selx,sely,x,y
        for i in range(TableN):
            q=list(Clistbox02.curselection())
            if x[i]==x[q[0]]:
                sely=i
        Clabel03.config(text=f"y축 : {x[sely]}")

    
    root02=Tk()
    root02.geometry("325x200")
    Clabel01=Label(root02,text="스케터 설정",width=16,
              font=("나눔고딕",TFontSize,"bold"),
              )
    Clabel02=Label(root02,text=f"x축 : {x[Selx]}",width=8,
              font=("나눔고딕",DFontSize,"bold"),
              )
    Clabel03=Label(root02,text=f"y축 : {x[Sely]}",width=8,
              font=("나눔고딕",DFontSize,"bold"),
              )
    Clistbox01=Listbox(root02,width=8,
                       font=("나눔고딕",DFontSize,"bold"),
                       )
    Clistbox02=Listbox(root02,width=8,
                       font=("나눔고딕",DFontSize,"bold"),
                       )
    for i in range(TableN):
        Clistbox01.insert(i,x[i])
        Clistbox02.insert(i,x[i])
    Clistbox01.bind('<<ListboxSelect>>', Sel1)
    Clistbox02.bind('<<ListboxSelect>>', Sel2)

    Clabel01.grid(row=0,column=0,columnspan=2,sticky=W+E+N+S)
    Clabel02.grid(row=1,column=0,sticky=W+E+N+S)
    Clabel03.grid(row=1,column=1,sticky=W+E+N+S)
    Clistbox01.grid(row=2,column=0,sticky=W+E+N+S)
    Clistbox02.grid(row=2,column=1,sticky=W+E+N+S)
    
    
    
#==============파일찾기==============#
FileLocate = filedialog.askopenfilename(initialdir="/",title = "excel 파일을 선택해 주세요")
FileName=ntpath.basename(FileLocate)
if FileLocate == '':
    messagebox.showwarning("경고", "파일을 찾을 수 없습니다.")    #파일 선택 안했을 때 메세지 출력
    root.destroy()
elif FileLocate[-5:]!='.xlsx':
    messagebox.showwarning("경고", "파일 확장자명이 xlsx가 아닙니다.")
    root.destroy()
wb = load_workbook(FileLocate, data_only=True)
ws = wb["Sheet1"]

TableN=ws.cell(row=2, column=2).value
NN=ws.cell(row=3,column=2).value
#==============정보추출==============#
List = [[0 for col in range(TableN)] for row in range(NN)]
for i in range(NN):
    for j in range(TableN):
        #a=ws.cell(row=i+2,column=j+4).value
        #print(f"{i} {j} : {a}")
        List[i][j]=ws.cell(row=i+2,column=j+4).value
        #List[i].append(a)
for i in range(TableN):
    a=ws.cell(row=1,column=i+4).value
    x.append(a)
for i in range(NN):
    b=ws.cell(row=i+2,column=3).value
    y.append(b)
xname=ws.cell(row=4,column=2).value
yname=ws.cell(row=5,column=2).value
unit=ws.cell(row=6,column=2).value
title=ws.cell(row=1,column=2).value
#=================UI=================#
plt.figure("IPSV")
label01=Label(root,text="IPSV 컨트롤러",width=16,
              font=("나눔고딕",TFontSize,"bold"),
              )
button01=Button(root,text="바형 그래프",width=8,
                font=("나눔고딕",DFontSize,"bold"),
                command=TypeB
             )
button02=Button(root,text="꺾은선 그래프",width=8,
              font=("나눔고딕",DFontSize,"bold"),
                command=TypeA
             )
button03=Button(root,text="스케터",width=8,
              font=("나눔고딕",DFontSize,"bold"),
                command=TypeC
             )
button04=Button(root,text="스케터 설정",width=8,
              font=("나눔고딕",DFontSize,"bold"),
                command=TypeC_2
             )

label01.grid(row=0,column=0,columnspan=2,sticky=W+E+N+S)
button01.grid(row=1,column=0,sticky=W+E+N+S)
button02.grid(row=1,column=1,sticky=W+E+N+S)
button03.grid(row=2,column=0,sticky=W+E+N+S)
button04.grid(row=2,column=1,sticky=W+E+N+S)

root.mainloop()

